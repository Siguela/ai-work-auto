import time
import os
import datetime
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.edge.options import Options

BASE_URL_1 = "https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c9280870193d0dcd0dc8f480193ed733b606832"
BASE_URL_2 = "https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019444dd44dd5dc901944a9a09e444c0"
USERNAME = ""
PASSWORD = ""
DOWNLOAD_DIR = r"C:\Users\wangxiaoyu02\工作项目\工作报表\trae分析项目\周报结论"

def get_timestamp_filename(original_name):
    """在文件名后添加年月日时间戳"""
    # 获取当前日期
    today = datetime.datetime.now()
    date_str = today.strftime("%Y%m%d")  # 格式：20260505
    
    # 分离文件名和扩展名
    name, ext = os.path.splitext(original_name)
    
    # 构建新文件名：原文件名 + 日期 + 扩展名
    new_name = f"{name}--{date_str}{ext}"
    return new_name

def analyze_report_and_add_conclusion(file_path):
    """分析报表数据并在Sheet2中添加结论"""
    try:
        # 用data_only=True打开文件，读取公式计算后的数值
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        sheet1 = workbook_data.active
        
        print("\n📋 报表结构分析：")
        print(f"- 表格行数: {sheet1.max_row}")
        print(f"- 表格列数: {sheet1.max_column}")
        
        team_data = {
            "海外团队": 6,
            "港澳": 7,
            "台湾": 8,
            "欧美澳": 9
        }
        
        overseas_team_rate = "%"
        europe_rate = "%"
        hk_macau_rate = "%"
        taiwan_rate = "%"
        
        print("- 数据提取详情:")
        for team_name, row_num in team_data.items():
            if row_num <= sheet1.max_row:
                cell_b = sheet1.cell(row=row_num, column=2).value
                cell_i = sheet1.cell(row=row_num, column=9).value
                
                print(f"  行{row_num}: B{row_num}='{cell_b}', I{row_num}='{cell_i}'")
                
                rate_str = "%"
                if cell_i is not None:
                    if isinstance(cell_i, (int, float)):
                        rate_str = f"{cell_i * 100:.2f}%"
                    else:
                        rate_str = str(cell_i)
                
                if team_name == "海外团队":
                    overseas_team_rate = rate_str
                elif team_name == "欧美澳":
                    europe_rate = rate_str
                elif team_name == "港澳":
                    hk_macau_rate = rate_str
                elif team_name == "台湾":
                    taiwan_rate = rate_str
        
        workbook_data.close()
        
        # 重新打开文件（不带data_only），以便写入Sheet2
        workbook = openpyxl.load_workbook(file_path)
        
        if 'Sheet2' in workbook.sheetnames:
            sheet2 = workbook['Sheet2']
        else:
            sheet2 = workbook.create_sheet('Sheet2')
        
        sheet2.delete_rows(1, sheet2.max_row)
        
        default_font = openpyxl.styles.Font(name='微软雅黑', size=11)
        
        conclusion_line1 = f"1）一单结课续费率：整体 {overseas_team_rate}（MTD目标 ，本月目标 ），欧美澳 {europe_rate}（MTD目标 ，本月目标 ），港澳 {hk_macau_rate}（MTD目标 ，本月目标 ），台湾 {taiwan_rate}（MTD目标 ，本月目标 ）；"
        
        sheet2['A1'] = conclusion_line1
        sheet2['A1'].font = default_font
        
        workbook.save(file_path)
        print(f"✓ 已添加分析结论到Sheet2")
        
        print("\n📊 分析结论预览：")
        print(f"  {conclusion_line1}")
        
        print("\n📈 数据提取结果：")
        print(f"  - 海外团队续费率: {overseas_team_rate}")
        print(f"  - 欧美澳续费率: {europe_rate}")
        print(f"  - 港澳续费率: {hk_macau_rate}")
        print(f"  - 台湾续费率: {taiwan_rate}")
        
        return True
        
    except Exception as e:
        print(f"⚠ 分析报表失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def analyze_report2_and_add_conclusion(file_path):
    """分析第二个报表数据并在Sheet2中添加结论"""
    try:
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        sheet1 = workbook_data.active
        
        print("\n📋 第二个报表结构分析：")
        print(f"- 表格行数: {sheet1.max_row}")
        print(f"- 表格列数: {sheet1.max_column}")
        
        # 在B列中搜索小组名称，从M列提取续费率
        target_teams = ["海外团队", "欧美澳", "港澳", "台湾"]
        team_rates = {}
        
        print("- 数据提取详情:")
        for row in range(6, min(30, sheet1.max_row + 1)):
            cell_b = sheet1.cell(row=row, column=2).value
            cell_m = sheet1.cell(row=row, column=13).value
            
            if cell_b:
                cell_b_str = str(cell_b).strip()
                for team in target_teams:
                    if team in cell_b_str and team not in team_rates:
                        rate_str = "%"
                        if cell_m is not None:
                            if isinstance(cell_m, (int, float)):
                                rate_str = f"{cell_m * 100:.2f}%"
                            else:
                                rate_str = str(cell_m)
                        team_rates[team] = rate_str
                        print(f"  行{row}: B{row}='{cell_b}', M{row}='{cell_m}' → {team}: {rate_str}")
        
        workbook_data.close()
        
        overseas_team_rate = team_rates.get("海外团队", "%")
        europe_rate = team_rates.get("欧美澳", "%")
        hk_macau_rate = team_rates.get("港澳", "%")
        taiwan_rate = team_rates.get("台湾", "%")
        
        workbook = openpyxl.load_workbook(file_path)
        
        if 'Sheet2' in workbook.sheetnames:
            sheet2 = workbook['Sheet2']
        else:
            sheet2 = workbook.create_sheet('Sheet2')
        
        sheet2.delete_rows(1, sheet2.max_row)
        
        default_font = openpyxl.styles.Font(name='微软雅黑', size=11)
        
        conclusion_line1 = f"2）统合结课续费率：整体 {overseas_team_rate}（MTD目标 ，本月目标 ），欧美澳 {europe_rate}（MTD目标 ，本月目标 ），港澳 {hk_macau_rate}（MTD目标 ，本月目标 ），台湾 {taiwan_rate}（MTD目标 ，本月目标 ）；"
        
        sheet2['A1'] = conclusion_line1
        sheet2['A1'].font = default_font
        
        workbook.save(file_path)
        print(f"✓ 已添加分析结论到Sheet2")
        
        print("\n📊 分析结论预览：")
        print(f"  {conclusion_line1}")
        
        print("\n📈 数据提取结果：")
        print(f"  - 海外团队续费率: {overseas_team_rate}")
        print(f"  - 欧美澳续费率: {europe_rate}")
        print(f"  - 港澳续费率: {hk_macau_rate}")
        print(f"  - 台湾续费率: {taiwan_rate}")
        
        return True
        
    except Exception as e:
        print(f"⚠ 分析第二个报表失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def js_export():
    print("使用JavaScript注入的导出脚本")
    
    # 创建下载目录
    if not os.path.exists(DOWNLOAD_DIR):
        os.makedirs(DOWNLOAD_DIR)
        print(f"✓ 创建下载目录: {DOWNLOAD_DIR}")
    
    edge_options = Options()
    edge_options.add_argument("--start-maximized")
    
    # 设置下载路径
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    edge_options.add_experimental_option("prefs", prefs)
    
    driver = webdriver.Edge(options=edge_options)
    
    try:
        # 步骤1: 打开页面
        print("步骤1: 打开页面")
        driver.get(BASE_URL_1)
        time.sleep(5)
        
        # 步骤2: 使用JavaScript登录
        print("\n步骤2: 使用JavaScript登录")
        login_script = f"""
            var username = document.querySelector('input[type="text"]');
            var password = document.querySelector('input[type="password"]');
            var loginBtn = document.querySelector('input.item-submit');
            
            if(username && password && loginBtn) {{
                username.value = '{USERNAME}';
                password.value = '{PASSWORD}';
                loginBtn.click();
                return '登录成功';
            }} else {{
                return '登录元素未找到';
            }}
        """
        result = driver.execute_script(login_script)
        print(f"  {result}")
        
        # 步骤3: 等待页面加载
        print("\n步骤3: 等待页面加载")
        for i in range(20):
            time.sleep(8)  # 减少等待时间
            try:
                page_source = driver.page_source
                if "btnExport" in page_source or "导出" in page_source:
                    print(f"  ✓ 页面加载完成 (第{i+1}次检查)")
                    break
            except Exception as e:
                print(f"  检查失败: {e}")
            print(f"  等待中... ({i+1}/20)")
        
        # 步骤4: 使用JavaScript点击导出按钮
        print("\n步骤4: 点击导出按钮")
        export_script = """
            var exportBtn = document.querySelector('input.btnExport');
            if(!exportBtn) exportBtn = document.querySelector('input[title="导出"]');
            if(!exportBtn) exportBtn = document.querySelector('input[value="导出"]');
            if(!exportBtn) exportBtn = document.querySelector('.queryview-toolbar-button');
            
            if(exportBtn) {
                exportBtn.click();
                console.log('导出按钮已点击');
                return '导出按钮已点击';
            } else {
                // 调试信息
                console.log('导出按钮未找到，页面上的按钮元素:', document.querySelectorAll('input, button'));
                return '导出按钮未找到';
            }
        """
        result = driver.execute_script(export_script)
        print(f"  {result}")
        
        # 等待导出弹窗加载
        print("  等待导出弹窗加载...")
        for i in range(10):
            time.sleep(3)
            hasExcel = driver.execute_script("return document.getElementById('EXCEL2007') !== null;")
            if hasExcel:
                print(f"  ✓ 导出弹窗已加载 (第{i+1}次检查)")
                break
            print(f"  等待弹窗... ({i+1}/10)")
        time.sleep(3)
        
        # 步骤5: 使用JavaScript选择Excel - 精确ID定位
        print("\n步骤5: 选择Excel")
        excel_script = """
            var foundExcel = null;
            var methodUsed = '';
            
            // 方式1: 优先使用ID定位（最可靠）
            var excelById = document.getElementById('EXCEL2007');
            if(excelById) {
                foundExcel = excelById;
                methodUsed = 'ID';
            }
            
            // 方式2: 使用caption属性定位
            if(!foundExcel) {
                var excelByCaption = document.querySelector('[caption="Excel"]');
                if(excelByCaption) {
                    foundExcel = excelByCaption;
                    methodUsed = 'caption属性';
                }
            }
            
            // 方式3: 使用class组合+文本定位
            if(!foundExcel) {
                var excelByClass = document.querySelector('div.dropdown-box-span-select.dropdown-box-span-higher');
                if(excelByClass && excelByClass.textContent && excelByClass.textContent.indexOf('Excel') !== -1) {
                    foundExcel = excelByClass;
                    methodUsed = 'class组合';
                }
            }
            
            // 方式4: 使用onmousedown事件特征定位
            if(!foundExcel) {
                var allDivs = document.querySelectorAll('div[onmousedown*="doItemClick"]');
                for(var i=0; i<allDivs.length; i++) {
                    if(allDivs[i].id && allDivs[i].id.indexOf('EXCEL') !== -1) {
                        foundExcel = allDivs[i];
                        methodUsed = 'onmousedown事件';
                        break;
                    }
                }
            }
            
            if(foundExcel) {
                // 调试信息：检查框架状态
                console.log('=== Excel按钮点击调试信息 ===');
                console.log('Excel按钮元素:', foundExcel);
                console.log('按钮ID:', foundExcel.id);
                console.log('按钮caption:', foundExcel.getAttribute('caption'));
                console.log('onmousedown函数:', typeof foundExcel.onmousedown);
                console.log('onmouseover函数:', typeof foundExcel.onmouseover);
                console.log('getOwner函数是否存在:', typeof getOwner === 'function');
                console.log('jsloader是否存在:', typeof jsloader !== 'undefined');
                if(window.jsloader) {
                    console.log('jsloader.resolve函数:', typeof jsloader.resolve);
                }
                
                var resultMessage = '';
                
                // 方法1: 优先使用框架方法
                try {
                    // 检查getOwner框架
                    if(typeof getOwner === 'function') {
                        var owner = getOwner(foundExcel);
                        if(owner && typeof owner.doItemClick === 'function') {
                            owner.doItemClick(foundExcel);
                            resultMessage = 'Excel已选择 (通过getOwner框架)';
                            console.log('✓ 通过getOwner框架成功');
                        }
                    }
                } catch(e) {
                    console.log('getOwner框架失败:', e.message);
                }
                
                // 方法2: 使用jsloader框架
                if(!resultMessage) {
                    try {
                        if(window.jsloader && typeof jsloader.resolve === 'function') {
                            var util = jsloader.resolve('freequery.common.util');
                            if(util && typeof util.getOwner === 'function') {
                                var owner = util.getOwner(foundExcel);
                                if(owner && typeof owner.doItemClick === 'function') {
                                    owner.doItemClick(foundExcel);
                                    resultMessage = 'Excel已选择 (通过jsloader框架)';
                                    console.log('✓ 通过jsloader框架成功');
                                }
                            }
                        }
                    } catch(e) {
                        console.log('jsloader框架失败:', e.message);
                    }
                }
                
                // 方法3: 触发原生事件
                if(!resultMessage) {
                    try {
                        // 先触发mouseover
                        if(typeof foundExcel.onmouseover === 'function') {
                            foundExcel.onmouseover();
                            console.log('✓ 触发onmouseover');
                        }
                        
                        // 再触发mousedown（主要事件）
                        if(typeof foundExcel.onmousedown === 'function') {
                            foundExcel.onmousedown();
                            resultMessage = 'Excel已选择 (通过onmousedown事件)';
                            console.log('✓ 通过onmousedown事件成功');
                        } else {
                            // 如果onmousedown不存在，直接click
                            foundExcel.click();
                            resultMessage = 'Excel已选择 (通过click事件)';
                            console.log('✓ 通过click事件成功');
                        }
                    } catch(e) {
                        console.log('原生事件失败:', e.message);
                    }
                }
                
                // 方法4: 模拟鼠标事件作为最终兜底
                if(!resultMessage) {
                    try {
                        // 创建完整的鼠标事件序列
                        var mouseoverEvent = new MouseEvent('mouseover', {
                            bubbles: true,
                            cancelable: true,
                            view: window
                        });
                        foundExcel.dispatchEvent(mouseoverEvent);
                        
                        var mousedownEvent = new MouseEvent('mousedown', {
                            bubbles: true,
                            cancelable: true,
                            view: window,
                            button: 0
                        });
                        foundExcel.dispatchEvent(mousedownEvent);
                        
                        var mouseupEvent = new MouseEvent('mouseup', {
                            bubbles: true,
                            cancelable: true,
                            view: window,
                            button: 0
                        });
                        foundExcel.dispatchEvent(mouseupEvent);
                        
                        var clickEvent = new MouseEvent('click', {
                            bubbles: true,
                            cancelable: true,
                            view: window
                        });
                        foundExcel.dispatchEvent(clickEvent);
                        
                        resultMessage = 'Excel已选择 (通过模拟事件)';
                        console.log('✓ 通过模拟事件成功');
                    } catch(e) {
                        console.log('模拟事件失败:', e.message);
                        resultMessage = '所有方法都失败: ' + e.message;
                    }
                }
                
                return resultMessage + ' (通过' + methodUsed + '定位)';
            } else {
                return 'Excel未找到';
            }
        """
        result = driver.execute_script(excel_script)
        print(f"  {result}")
        time.sleep(10)
        
        # 步骤6: 使用JavaScript点击在线导出 - 精确定位
        print("\n步骤6: 点击在线导出")
        online_export_script = """
            var foundOnline = null;
            
            // 方式1: 使用class和value精确匹配
            var onlineByClass = document.querySelector('input._btnOK.barbtn.btn-default[value="在线导出"]');
            if(onlineByClass) {
                foundOnline = onlineByClass;
                console.log('✓ 通过class和value找到在线导出按钮');
            }
            
            // 方式2: 使用qtp属性定位
            if(!foundOnline) {
                var onlineByQtp = document.querySelector('[qtp="baseDialogEx_btnOK"]');
                if(onlineByQtp) {
                    foundOnline = onlineByQtp;
                    console.log('✓ 通过qtp属性找到在线导出按钮');
                }
            }
            
            // 方式3: 使用class组合定位
            if(!foundOnline) {
                var onlineByClassOnly = document.querySelector('input._btnOK.barbtn.btn-default');
                if(onlineByClassOnly && onlineByClassOnly.value === '在线导出') {
                    foundOnline = onlineByClassOnly;
                    console.log('✓ 通过class组合找到在线导出按钮');
                }
            }
            
            // 方式4: 文本匹配作为兜底
            if(!foundOnline) {
                var allInputs = document.querySelectorAll('input[type="button"]');
                for(var i=0; i<allInputs.length; i++) {
                    if(allInputs[i].value && allInputs[i].value.trim() === '在线导出') {
                        foundOnline = allInputs[i];
                        console.log('✓ 通过文本匹配找到在线导出按钮');
                        break;
                    }
                }
            }
            
            if(foundOnline) {
                // 调试信息
                console.log('在线导出按钮元素:', foundOnline);
                console.log('按钮class:', foundOnline.className);
                console.log('按钮value:', foundOnline.value);
                console.log('按钮qtp:', foundOnline.getAttribute('qtp'));
                
                // 点击按钮
                foundOnline.click();
                return '在线导出已点击';
            } else {
                // 调试信息
                console.log('未找到在线导出按钮，页面上的按钮元素:');
                var allButtons = document.querySelectorAll('input[type="button"], button');
                for(var j=0; j<allButtons.length; j++) {
                    console.log('按钮' + j + ':', allButtons[j].value || allButtons[j].textContent, allButtons[j].className);
                }
                return '在线导出未找到';
            }
        """
        result = driver.execute_script(online_export_script)
        print(f"  {result}")
        time.sleep(30)
        
        print("\n✅ 导出完成！")
        
        # 检查并重命名最新下载的文件
        print("\n检查下载的文件...")
        time.sleep(10)  # 给文件下载一些时间
        
        # 获取下载目录中的文件
        if os.path.exists(DOWNLOAD_DIR):
            # 获取所有文件及其创建时间
            files_info = []
            for file in os.listdir(DOWNLOAD_DIR):
                file_path = os.path.join(DOWNLOAD_DIR, file)
                if os.path.isfile(file_path):
                    create_time = os.path.getctime(file_path)
                    files_info.append((file, file_path, create_time))
            
            if files_info:
                # 按创建时间排序，获取最新文件
                files_info.sort(key=lambda x: x[2], reverse=True)
                latest_file, latest_path, latest_time = files_info[0]
                
                print(f"✓ 下载目录中共有 {len(files_info)} 个文件")
                print(f"✓ 最新文件: {latest_file} (创建时间: {datetime.datetime.fromtimestamp(latest_time)})")
                
                # 只重命名最新文件
                file_size = os.path.getsize(latest_path)
                new_filename = get_timestamp_filename(latest_file)
                new_filepath = os.path.join(DOWNLOAD_DIR, new_filename)
                
                try:
                    os.rename(latest_path, new_filepath)
                    print(f"✓ 重命名最新文件: {latest_file} → {new_filename} ({file_size} 字节)")
                    print(f"✅ 文件已保存为: {new_filename}")
                    
                    # 分析报表并添加结论
                    print("\n开始分析报表数据...")
                    if analyze_report_and_add_conclusion(new_filepath):
                        print("✅ 报表分析完成！")
                    else:
                        print("⚠ 报表分析失败，但文件已保存")
                        
                except Exception as e:
                    print(f"⚠ 重命名失败: {latest_file} → {e}")
                    print(f"⚠ 文件保持原名: {latest_file}")
            else:
                print("⚠ 下载目录为空，可能下载失败")
        else:
            print("❌ 下载目录不存在")
        
        # ==================== 第二个报表导出 ====================
        print("\n" + "=" * 50)
        print("开始导出第二个报表...")
        print("=" * 50)
        
        # 步骤7: 打开第二个报表页面
        print("\n步骤7: 打开第二个报表页面")
        driver.get(BASE_URL_2)
        time.sleep(5)
        
        # 步骤8: 等待页面加载
        print("\n步骤8: 等待页面加载")
        for i in range(20):
            time.sleep(8)
            try:
                page_source = driver.page_source
                if "btnExport" in page_source or "导出" in page_source:
                    print(f"  ✓ 页面加载完成 (第{i+1}次检查)")
                    break
            except Exception as e:
                print(f"  检查失败: {e}")
            print(f"  等待中... ({i+1}/20)")
        
        # 步骤9: 点击导出按钮
        print("\n步骤9: 点击导出按钮")
        result = driver.execute_script(export_script)
        print(f"  {result}")
        
        # 等待导出弹窗加载
        print("  等待导出弹窗加载...")
        for i in range(10):
            time.sleep(3)
            hasExcel = driver.execute_script("return document.getElementById('EXCEL2007') !== null;")
            if hasExcel:
                print(f"  ✓ 导出弹窗已加载 (第{i+1}次检查)")
                break
            print(f"  等待弹窗... ({i+1}/10)")
        time.sleep(3)
        
        # 步骤10: 选择Excel
        print("\n步骤10: 选择Excel")
        result = driver.execute_script(excel_script)
        print(f"  {result}")
        time.sleep(10)
        
        # 步骤11: 点击在线导出
        print("\n步骤11: 点击在线导出")
        result = driver.execute_script(online_export_script)
        print(f"  {result}")
        time.sleep(30)
        
        print("\n✅ 第二个报表导出完成！")
        
        # 检查并重命名最新下载的文件
        print("\n检查第二个报表下载的文件...")
        time.sleep(10)
        
        if os.path.exists(DOWNLOAD_DIR):
            files_info = []
            for file in os.listdir(DOWNLOAD_DIR):
                file_path = os.path.join(DOWNLOAD_DIR, file)
                if os.path.isfile(file_path):
                    create_time = os.path.getctime(file_path)
                    files_info.append((file, file_path, create_time))
            
            if files_info:
                files_info.sort(key=lambda x: x[2], reverse=True)
                latest_file, latest_path, latest_time = files_info[0]
                
                # 检查是否是刚下载的新文件（创建时间在最近2分钟内）
                time_diff = time.time() - latest_time
                if time_diff > 120:
                    print("⚠ 未检测到新下载的文件")
                else:
                    print(f"✓ 最新文件: {latest_file} (创建时间: {datetime.datetime.fromtimestamp(latest_time)})")
                    
                    file_size = os.path.getsize(latest_path)
                    new_filename = get_timestamp_filename(latest_file)
                    new_filepath = os.path.join(DOWNLOAD_DIR, new_filename)
                    
                    try:
                        os.rename(latest_path, new_filepath)
                        print(f"✓ 重命名: {latest_file} → {new_filename} ({file_size} 字节)")
                        print(f"✅ 第二个报表已保存为: {new_filename}")
                        
                        print("\n开始分析第二个报表数据...")
                        if analyze_report2_and_add_conclusion(new_filepath):
                            print("✅ 第二个报表分析完成！")
                        else:
                            print("⚠ 第二个报表分析失败，但文件已保存")
                            
                    except Exception as e:
                        print(f"⚠ 重命名失败: {latest_file} → {e}")
            else:
                print("⚠ 下载目录为空，可能下载失败")
        else:
            print("❌ 下载目录不存在")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        time.sleep(10)
        driver.quit()

if __name__ == "__main__":
    js_export()
